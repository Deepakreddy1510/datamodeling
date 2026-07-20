import datetime
<<<<<<< HEAD
import json
from decimal import Decimal
from enum import Enum
from uuid import UUID
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROLE_COLORS = {
    "raw": "C65911",
    "staging": "0F6B63",
    "dimension": "7030A0",
    "fact": "1F4E78",
    "other": "595959",
}
LIGHT_FILLS = {
    "raw": "FCE4D6",
    "staging": "DDEBF7",
    "dimension": "E4DFEC",
    "fact": "D9EAF7",
    "other": "F2F2F2",
}


def _excel_safe_value(value):
    """Convert PostgreSQL/Python values into safe openpyxl cell values."""
    if value is None:
        return None

    if isinstance(value, UUID):
        return str(value)

    if isinstance(value, Enum):
        value = value.value

    if isinstance(value, datetime.datetime):
        if value.tzinfo is not None:
            return value.astimezone(datetime.timezone.utc).replace(tzinfo=None)
=======
from pathlib import Path

from openpyxl import Workbook


def _excel_safe_value(value):
    """Convert Python values into Excel-safe values."""
    if isinstance(value, datetime.datetime):
        if value.tzinfo is not None:
            return value.replace(tzinfo=None)
>>>>>>> personal/main
        return value

    if isinstance(value, datetime.time):
        if value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value

<<<<<<< HEAD
    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (bytes, bytearray, memoryview)):
        return bytes(value).hex()

    if isinstance(value, (dict, list, tuple, set)):
        serializable = list(value) if isinstance(value, set) else value
        return json.dumps(serializable, ensure_ascii=False, sort_keys=True, default=str)

    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        # Prevent generated text from being interpreted as an Excel formula.
        return "'" + value

    return value


def _table_role(table_name):
    name = table_name.lower()
    if name.startswith("load_") or name.endswith("_raw"):
        return "raw"
    if name.startswith("stg_"):
        return "staging"
    if name.startswith("dim_"):
        return "dimension"
    if name.startswith("fact_"):
        return "fact"
    return "other"


def _purpose(role):
    return {
        "raw": "Source-aligned ingestion records and original payloads",
        "staging": "Typed, cleansed, transformation-ready records",
        "dimension": "Conformed descriptive entities and surrogate keys",
        "fact": "Transactional or periodic analytical measures",
        "other": "Supporting model table",
    }[role]


def _style_table_sheet(worksheet, table, row_count):
    role = _table_role(table.name)
    color = ROLE_COLORS[role]
    light = LIGHT_FILLS[role]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 24

    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Subtle banding without creating Excel table-name collisions.
    for row_index in range(2, min(row_count + 2, worksheet.max_row + 1)):
        if row_index % 2 == 0:
            for cell in worksheet[row_index]:
                cell.fill = PatternFill("solid", fgColor=light)

    for column_index, column in enumerate(table.columns, start=1):
        letter = get_column_letter(column_index)
        max_length = len(column.name)
        for row_index in range(2, min(worksheet.max_row, 202) + 1):
            value = worksheet.cell(row_index, column_index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        is_payload = column.name.lower() in {"source_payload", "payload", "raw_payload"}
        width = 55 if is_payload else min(max(max_length + 2, 12), 36)
        worksheet.column_dimensions[letter].width = width

        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row_index, column_index)
            if isinstance(cell.value, datetime.datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif isinstance(cell.value, datetime.date):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, datetime.time):
                cell.number_format = "hh:mm:ss"
            if is_payload or (isinstance(cell.value, str) and len(cell.value) > 50):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")


def _write_summary(workbook, model, data):
    worksheet = workbook.create_sheet(title="Model_Summary", index=0)
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:F2")
    worksheet["A1"] = "Synthetic Data Warehouse — Model Summary"
    worksheet["A1"].fill = PatternFill("solid", fgColor="17365D")
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Layer", "Table", "Rows", "Columns", "Purpose", "Status"]
    worksheet.append([])
    worksheet.append(headers)
    for cell in worksheet[4]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for table in model.tables:
        role = _table_role(table.name)
        worksheet.append([
            role.title(),
            table.name,
            len(data.get(table.name, [])),
            len(table.columns),
            _purpose(role),
            "Generated",
        ])
        row = worksheet.max_row
        fill = PatternFill("solid", fgColor=LIGHT_FILLS[role])
        for cell in worksheet[row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = f"A4:F{worksheet.max_row}"
    widths = [14, 30, 12, 12, 58, 14]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    return worksheet


=======
    return value


>>>>>>> personal/main
def write_excel(model, data, output_path):
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
<<<<<<< HEAD
    _write_summary(workbook, model, data)
=======
>>>>>>> personal/main

    for table in model.tables:
        worksheet = workbook.create_sheet(title=table.name[:31])
        headers = [column.name for column in table.columns]
        worksheet.append(headers)

<<<<<<< HEAD
        rows = data.get(table.name, [])
        for row in rows:
            worksheet.append(
                [_excel_safe_value(row.get(column.name)) for column in table.columns]
            )
        _style_table_sheet(worksheet, table, len(rows))

    workbook.save(path)
=======
        for row in data.get(table.name, []):
            worksheet.append(
                [_excel_safe_value(row.get(column.name)) for column in table.columns]
            )

    workbook.save(path)
>>>>>>> personal/main
